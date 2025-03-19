import json
from openpyxl import load_workbook
import sys

def transfer_anonymized_values(source_file, anonymized_file, mapping_file, output_file):
    # Ladda mappningsfilen
    with open(mapping_file, 'r', encoding='utf-8') as f:
        mappings = json.load(f)
    
    # Öppna både källdokumentet och det anonymiserade dokumentet
    source_wb = load_workbook(source_file)
    anonymized_wb = load_workbook(anonymized_file)
    
    # Antag att vi arbetar med första arbetsbladet i båda dokumenten
    source_ws = source_wb.active
    anonymized_ws = anonymized_wb.active
    
    # Skapa en omvänd mappning för enklare uppslagning
    reverse_mappings = {}
    for mapping_type in mappings.values():
        for original, anonymized in mapping_type.items():
            reverse_mappings[anonymized] = original
    
    # Gå igenom alla celler i det anonymiserade dokumentet
    updated_cells = 0
    for row in range(1, anonymized_ws.max_row + 1):
        for col in range(1, anonymized_ws.max_column + 1):
            anon_cell = anonymized_ws.cell(row=row, column=col)
            if anon_cell.value in reverse_mappings:
                # Hitta motsvarande cell i källdokumentet
                source_cell = source_ws.cell(row=row, column=col)
                # Spara originalvärdet för loggning
                original_value = source_cell.value
                # Uppdatera endast värdet, behåll all formatering
                source_cell.value = anon_cell.value
                updated_cells += 1
                print(f"Uppdaterade cell {source_ws.cell(row=row, column=1).column_letter}{row}: "
                      f"'{original_value}' -> '{anon_cell.value}'")
    
    # Spara det uppdaterade dokumentet med nytt namn
    source_wb.save(output_file)
    print(f"\nKlart! {updated_cells} celler har uppdaterats.")
    print(f"Det formaterade resultatet har sparats i: {output_file}")

if __name__ == "__main__":
    if len(sys.argv) != 5:
        print("Användning: python preserve_formatting.py källfil.xlsx anonymiserad.xlsx mappning.json utfil.xlsx")
        sys.exit(1)
    
    source_file = sys.argv[1]
    anonymized_file = sys.argv[2]
    mapping_file = sys.argv[3]
    output_file = sys.argv[4]
    
    transfer_anonymized_values(source_file, anonymized_file, mapping_file, output_file) 